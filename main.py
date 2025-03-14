# !/usr/bin/env python
# coding: utf-8

import os

from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

import time
from lxml import etree

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException,StaleElementReferenceException,ElementClickInterceptedException

from LOGGER import GetLogger

logger = GetLogger(logger_name="XiaoHongShu", debug=False, log_file="XiaoHongShu.log")

# 设置 Chrome 选项以避免打印出过多日志
chrome_options = Options()
USERPATH = os.path.abspath("./userData")
service = Service('/Users/737chen/Downloads/chromedriver-mac-arm64/chromedriver')
chrome_options.add_argument(f"--user-data-dir={USERPATH}")
chrome_options.add_argument("--headless=new")  # 无头模式
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--window-size=1920,1080")
# chrome_options.add_argument("--window-size=1280,720")


chrome_options.add_argument("--disable-logging")
chrome_options.add_argument("--log-level=3")
# 浏览器伪装:模拟真实的用户在操作.防反爬
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
chrome_options.add_experimental_option("useAutomationExtension", False)
# 初始化 WebDriver
driver = webdriver.Chrome(options=chrome_options,service=service)

LOGIN_CONTAINER_CLASS = 'login-container'
LOGIN_BTN_CLASS = 'login-btn'
USER_BTN_CLASS = 'user'

XIAOHONGSHU_URL = 'https://www.xiaohongshu.com/explore'
        
def search_page(keyword,total):
    logger.info("打开小红书页面中")
    driver.get(XIAOHONGSHU_URL)
    SEARCH_INPUT_ID = 'search-input'
    SEARCH_ICON_CLASS = 'search-icon'
    # 向搜索框发送关键字
    logger.info(f"搜索关键字[{keyword}]中")
    search_input_object = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, SEARCH_INPUT_ID)))
    search_input_object.send_keys(keyword)
    logger.info(f"输入关键字[{keyword}]完成")
    # 点击搜索按钮
    logger.info(f"点击搜索按钮中")
    search_input_button_object = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, SEARCH_ICON_CLASS)))
    search_input_button_object.click()
    logger.info(f"点击搜索按钮完成")
    time.sleep(2)
    # 点击筛选按钮
    FILTER_INCO_CLASS = 'filter'
    filter_area_object = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CLASS_NAME, FILTER_INCO_CLASS)))
    # 悬浮至该按钮上
    ActionChains(driver).move_to_element(filter_area_object).perform()
    time.sleep(1)  # 等待下拉菜单显示
    # 获取所有具有类名 'dropdown-items' 的元素
    dropdown_menus = driver.find_elements(By.CLASS_NAME, "dropdown-items")

    # 过滤出 display 样式为 inline-block 的元素（就是正常显示的那个）
    visible_menus = [
        menu for menu in dropdown_menus 
        if driver.execute_script("return window.getComputedStyle(arguments[0]).display;", menu) == 'inline-block'
    ]
    # 输出符合条件的元素数量及内容
    logger.info(f"找到 {len(visible_menus)} 个 display 为 inline-block 的 'dropdown-items' 元素")
        
    if len(visible_menus) == 0:
        logger.error("未找到任何可见的菜单")
        return
    
    # 获取第一个可见的菜单
    dropdown_menu = visible_menus[0]
    # 获取第一个可见的菜单中的所有 <li> 元素
    li_elements = dropdown_menu.find_elements(By.TAG_NAME, "li")
    # 遍历每个 <li>，查找其中包含文本 "最热" 的 <span>
    for li in li_elements:
        try:
            # 查找 <li> 内的 <span> 标签
            span = li.find_element(By.TAG_NAME, "span")
            # 检查 <span> 的文本是否为 "最热"
            if span.text == "最热":
                # 点击找到的 "最热" 选项
                li.click()
                logger.info("成功点击 '最热' 选项")
                break
        except Exception as e:
            logger.warning(f"在<li>元素中查找<span>时出现问题: {e}")
    else:
        logger.error("未找到包含 '最热' 文本的 <li> 标签")


    last_index = 0 # 上一次最后爬取的数据索引
    no_change_count = 0  # 记录没有变化的次数
    no_change_limit = 3  # 设定限制次数
    while last_index <= int(total - 1):
        if last_index == 0:
            logger.info("开始爬取数据中")
        else:
            logger.info(f"该翻页了! 起始数据{last_index + 1}条: 获取页面数据中")
        # 获取内容页面
        FEEDS_CONTAINER_CLASS = 'feeds-container'
        feed_container_object = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, FEEDS_CONTAINER_CLASS)))
        # 获取容器内的所有 <section> 标签（代表每个帖子）
        sections = feed_container_object.find_elements(By.TAG_NAME, "section")
        # 检测数据变化的循环
        previous_last_index = last_index
        # 如果已经爬取了数据则需要对页面进行滚动加载操作
        # 由于数据不是一条条加载而是一次加载数个的所以要不断尝试
        if last_index != 0:
            logger.info(f"开始滚动页面")
            ActionChains(driver).scroll_by_amount(0, 600).perform()
            time.sleep(3)
            feed_container_object = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, FEEDS_CONTAINER_CLASS)))
            sections = feed_container_object.find_elements(By.TAG_NAME, "section")
            try:
                # 获取第一个和最后一个 <section> 标签的数据索引
                first_section = sections[0].get_attribute("data-index")
                last_section = sections[-1].get_attribute("data-index")
            except StaleElementReferenceException:
                logger.warning("检测到 StaleElementReferenceException，重新获取 sections")
                sections = feed_container_object.find_elements(By.TAG_NAME, "section")
                first_section = sections[0].get_attribute("data-index")
                last_section = sections[-1].get_attribute("data-index")
            
            while int(first_section) > int(last_index):
                logger.warning(f"滑动超过预期,回滑")
                ActionChains(driver).scroll_by_amount(0, -200).perform()
                feed_container_object = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, FEEDS_CONTAINER_CLASS)))
                sections = feed_container_object.find_elements(By.TAG_NAME, "section")
                first_section = sections[0].get_attribute("data-index")
                last_section = sections[-1].get_attribute("data-index")
            
            logger.info(f"回滑完成,当前数据开始索引为{first_section},结束索引为{last_section}")
            last_index = get_container(keywords,total,sections,last_index)
        else:
            last_index = get_container(keywords,total,sections,last_index)
            
        
        # 检查是否数据变化
        if last_index == previous_last_index:
            no_change_count += 1
            if no_change_count >= no_change_limit:
                logger.info("数据已停止变化，结束爬取")
                break
        else:
            no_change_count = 0  # 重置计数器


def get_container(keywords,total,sections,last_index):
    file  = f"{keywords}.xlsx"
    # 获取内容页面
    FEEDS_CONTAINER_CLASS = 'feeds-container'
    feed_container_object = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, FEEDS_CONTAINER_CLASS)))
    time.sleep(2)
    logger.info(f"获取页面中卡片完成")

    # 动态获取刷新以对抗爬虫过程中页面元素更新
    for i in range(len(sections)):
        if last_index == total:
            logger.info(f"所有帖子爬取完毕")
            break
        try:
            logger.info(f"读取第{last_index}个帖子")
            # 重新定位当前 section，以防止 stale element 问题
            sections = feed_container_object.find_elements(By.TAG_NAME, "section")
            # 判断是否包含 data-width 和 data-height 属性，对抗广告
            data_index = sections[i].get_attribute("data-index")
            data_width = sections[i].get_attribute("data-width")
            data_height = sections[i].get_attribute("data-height")
            if int(data_index) < int(last_index):
                logger.info(f"本批次, 第 {i} 个帖子爬取了，跳过")
                continue
            if not data_width or not data_height:
                logger.info(f"第 {last_index} 个帖子为广告，跳过")
                last_index += 1
                tree = None
            else:
                logger.info(f"点击第{last_index}个帖子")
                try:
                    sections[i].click()
                except ElementClickInterceptedException:
                    logger.error(f"第 {last_index} 个帖子点击失败,可能是不兼容的直播或广告,跳过")
                    tree = None
                    last_index += 1
                    get_content(tree,file) # 获取页面数据
                    continue
                # 抓取帖子页面内容
                AUTHOR_AREA_CLASS = 'author-wrapper'
                TITLE_AREA_ID = 'detail-title'
                CONTENT_AREA_ID = 'detail-desc'
                BOTTOM_AREA_CLASS = 'bottom-container'
                INTERATE_AREA_CLASS = 'interact-container'
                logger.info("加载数据中")
                wait = WebDriverWait(driver, 10)
                # 等待页面的各个部分加载完成
                wait.until(EC.presence_of_element_located((By.CLASS_NAME, AUTHOR_AREA_CLASS)))
                logger.info("加载作者信息完成")
                try:
                    wait.until(EC.presence_of_element_located((By.ID, TITLE_AREA_ID)))
                    logger.info("加载标题完成")
                    wait.until(EC.presence_of_element_located((By.ID, CONTENT_AREA_ID)))
                    logger.info("加载内容完成")
                    wait.until(EC.presence_of_element_located((By.CLASS_NAME, BOTTOM_AREA_CLASS)))
                    logger.info("加载底部信息完成")
                except TimeoutException:
                    logger.warning("出现内容为空,此消息作为警告消息提示")
                except Exception as e:
                    logger.error(f"加载内容失败，错误信息：{e}")
                
                wait.until(EC.presence_of_element_located((By.CLASS_NAME, INTERATE_AREA_CLASS)))
                logger.info("加载具体浏览数据完成")
                # 截取页面数据
                logger.info("截取当前页面数据中...")
                page_content = driver.page_source
                tree = etree.HTML(page_content)
                
                time.sleep(2)
                # 发送 ESC 键关闭帖子
                logger.info("点击关闭按钮关闭帖子")
                CLOSS_BTN_CLASS = 'close-circle'
                close_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, CLOSS_BTN_CLASS)))
                close_btn.click()
                logger.info("已关闭帖子")

                # 更新读取进度
                logger.info(f"读取第{last_index}个帖子完成")
                last_index += 1
                # 等待关闭后重新定位页面
                logger.info("重新定位页面中")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, FEEDS_CONTAINER_CLASS)))
                logger.info("重新定位页面完成")
                time.sleep(3)
                
            get_content(tree,file) # 获取页面数据
        except StaleElementReferenceException:
            logger.error(f"StaleElementReferenceException: 第{last_index}个帖子元素失效，重新尝试获取元素")
            # 重新获取页面的帖子列表并重试
            feed_container_object = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, FEEDS_CONTAINER_CLASS)))
            sections = feed_container_object.find_elements(By.TAG_NAME, "section")
        except IndexError:
            logger.warning(f"已经爬取到了所有帖子无法再爬取更多了，总数为{last_index}个")
    logger.info(f"爬取完成,爬取到{last_index}个帖子")
    return last_index

def get_content(tree,file):
    data = {}
    if not tree:
        data["remark"] = "这是一个广告"
        write_to_excel(data,file)
        return

    # 获取用户名
    username_element = tree.xpath('//div[@class="author-wrapper"]//span[@class="username"]/text()')
    if username_element:
        data["username"] = username_element[0]
    
    # 获取标题
    title_element = tree.xpath('//div[@id="detail-title"]/text()')
    if title_element:
        data["title"] = title_element[0]

    # 获取内容
    content_element = tree.xpath('//div[@id="detail-desc"]//span[@class="note-text"]/span/text()')
    if content_element:
        data["content"] = content_element[0]


    '''
     # 获取标签
    tags_elements = tree.xpath('//a[@class="tag"]/text()')
    tags = [tag.strip() for tag in tags_elements]
    data["tags"] = tags

    # 获取发布时间和地点
    date_local_element = tree.xpath('//div[@class="bottom-container"]//span[@class="date"]/text()')
    if date_local_element:
        data["date_local"] = date_local_element[0]
    '''


    
    # 获取点赞数
    like_count_element = tree.xpath('//div[@class="interact-container"]/div/div//span[contains(@class, "like-wrapper")]//span[contains(@class, "count")]/text()')
    if like_count_element:
        data["like_count"] = like_count_element[0]
        if data["like_count"] == "点赞":
            data["like_count"] = 0
            
            
    # 获取收藏数
    collect_count_element = tree.xpath('//span[contains(@class, "collect-wrapper")]//span[contains(@class, "count")]/text()')
    if collect_count_element:
        data["collect_count"] = collect_count_element[0]
        if data["collect_count"] == "收藏":
            data["collect_count"] = 0

    # 获取评论数
    comment_count_element = tree.xpath('//span[contains(@class, "chat-wrapper")]//span[contains(@class, "count")]/text()')
    if comment_count_element:
        data["comment_count"] = comment_count_element[0]
        if data["comment_count"] == "评论":
            data["comment_count"] = 0

    # 获取一级评论---还需要新增：在页面内滑动的代码
    # first_floor_comment_element = tree.xpath('//span[@class="note-text"]/span/text()')
    first_floor_comment_element = tree.xpath('//div[@class="content"]/span[@class="note-text"]/span/text()')
    if first_floor_comment_element:
        print("first_floor_comment_element:", first_floor_comment_element)
        data["first_floor_comment"] = "\r\n".join(first_floor_comment_element)
    else:
        print("none of first_floor_comment")

    # 获取二级评论---还需要新增：点击「展开」，在评论span内滑动，分支判断是否需要多次点击「展开」/滑动页面


    # 将数据写入 Excel
    write_to_excel(data,file)
    
def write_to_excel(data, filename="output.xlsx"):
    try:
        # 尝试打开已有文件，否则创建新文件
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        # 写入表头
        headers = ["用户名", "标题", "内容",
                   #"标签", "发布时间和地点",
                   "点赞数", "收藏数", "评论数", "一级&二级评论"]
        ws.append(headers)
        # 设置表头样式
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 15  # 设置列宽
        
    
    # 写入数据到表格
    ws.append([
        data.get("username", ""),
        data.get("title", ""),
        data.get("content", ""),
        # ", ".join(data.get("tags", [])),
        # data.get("date_local", ""),
        data.get("like_count", ""),
        data.get("collect_count", ""),
        data.get("comment_count", ""),
        data.get("first_floor_comment", "")

    ])
    
    # 保存文件
    wb.save(filename)
    logger.info(f"数据已写入 {filename}")

if __name__ == '__main__':
    
    # 定义需要爬取的关键词
    keywords = "wps替代"
    total = 229 # 设置爬取的最大条数(小红书的最新页面展示的是229个)
    # 爬取关键词
    search_page(keywords,total)
